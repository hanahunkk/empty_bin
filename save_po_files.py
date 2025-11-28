import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font

columns_to_export = [
    "No", "Palllet#", "Tfc Code", "Preferred Bin",
    "Qt", "Target Bin1", "Target Bin2", "Memo"
]


def save_file(input_df, po_files_names: list[str]):

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    font_normal = Font(name="Arial", size=12)
    font_bold = Font(name="Arial", size=12, bold=True)

    po_files = [os.path.basename(p) for p in po_files_names]

    for po_name, df_group in input_df.groupby("PO name"):
        if pd.isna(po_name) or str(po_name).strip() == "":
            continue

        if po_name not in po_files:
            continue

        # print(f"Processing: {po_name}")
        # print(df_group)
        # print(f"df_group before sort:\n{df_group}")

        first_value = df_group["PO name"].iloc[0]

        df_group = df_group.sort_values(
            by=[ "Preferred Bin", "Palllet#" ],
            ascending=[True, True]
        ).reset_index(drop=True)
        merged_rows = []

        for (pallet, tfc), group in df_group.groupby(["Palllet#", "Tfc Code"]):
            record = group.iloc[0].copy()

            bins1 = group["Target Bin1"].dropna().unique().tolist()
            bins2 = group["Target Bin2"].dropna().unique().tolist()

            record["Target Bin1"] = "".join(bins1)
            record["Target Bin2"] = "".join(bins2)

            merged_rows.append(record)

        df_merged = pd.DataFrame(merged_rows)
        df_merged = df_merged.sort_values(
            by=["Preferred Bin", "Palllet#" ],
            ascending=[True, True]
        ).reset_index(drop=True)

        if "No" not in df_merged.columns:
            df_merged.insert(0, "No", range(1, len(df_merged) + 1))
        else:
            df_merged["No"] = range(1, len(df_merged) + 1)
        # print(f"po_name sheet name1: {po_name}")
        safe_name = re.sub(r'[\\/:\*\?"<>\|\-]', "_", po_name)
        safe_name = os.path.splitext(safe_name)[0]
        # print(f"safe_name sheet name1: {safe_name}")

        match = re.search(r"PO\d{5}([^_]+)", safe_name)
        if match:
            extracted = match.group(1)  # ✅ MM01AAD25
        else:
            extracted = po_name
        # print(f"Extracted sheet name1: {extracted}")
        extracted = re.sub(r'[\\/\:\*\?\[\]]', "_", extracted)
        extracted = extracted.strip()
        if len(extracted) > 31:
            extracted = extracted[:31]
        # print(f"Extracted sheet name2: {extracted}")

        output_path = os.path.join(f"{safe_name}_auto")

        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        df_merged[columns_to_export].to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        ws.insert_rows(1)
        ws.merge_cells("A1:H1")
        ws["A1"] = po_name
        ws["A1"].font = font_bold

        ws.title = extracted

        # Header fill
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = font_bold
            cell.border = thin_border

        # Cell line
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.font = font_normal
                cell.border = thin_border

        # Auto column width
        # for column_cells in ws.columns:
        #     max_length = 0
        #     column = column_cells[0].column_letter
        #     column_header = column_cells[0].value
        #
        #     for cell in column_cells:
        for col in ws.iter_cols(min_row=2):  # 🔥 2행부터 시작
            max_length = 0
            column_letter = col[0].column_letter  # 이제 정상 Cell

            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass

            adjusted_width = max_length + 2

            if col[0].value in ["Palllet#", "memo", "preferred bin", "Target Bin1", "Target Bin2"]:
                adjusted_width = 20

            ws.column_dimensions[column_letter].width = adjusted_width


        wb.save(output_path)
        wb.close()

        print(f"✅ Saved: {output_path}")


import os
import re

def extract_po_name(filepath):
    filename = os.path.basename(filepath)
    match = re.search(r'(PO\d{5}[A-Za-z0-9]+)', filename)
    return match.group(1) if match else None

