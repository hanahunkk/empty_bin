import os
import re
import shutil
import datetime
import pandas as pd
from tkinter import Tk, Button, messagebox
from openpyxl import load_workbook

SOURCE_DIR = r"\\Tfc-akl-share\tfc共有\物流部\物流部\ItemStockSearch"
BASE_TEMPLATE = "bins.xlsx"
TARGET_DIR = os.path.dirname(os.path.abspath(__file__))


def find_latest_stocklist(source_dir):
    pattern = re.compile(r"(\d+)-ItemStockList\.xlsx$")
    max_num = -1
    latest_file = None
    for file in os.listdir(source_dir):
        match = pattern.match(file)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
                latest_file = os.path.join(source_dir, file)
    return latest_file


def create_po_file():
    now = datetime.datetime.now()
    date_time = now.strftime("%Y%m%d%H%M")

    if not os.path.exists(BASE_TEMPLATE):
        messagebox.showerror("Error", f"Cannot find the Base file '{BASE_TEMPLATE}'.")
        return

    new_filename = f"bin_{date_time}.xlsx"
    target_path = os.path.join(TARGET_DIR, new_filename)
    shutil.copy2(BASE_TEMPLATE, target_path)

    source_file = find_latest_stocklist(SOURCE_DIR)
    if not source_file:
        messagebox.showerror("Error", "Cannot find the file 'ItemStockList'.")
        return

    wb_src = load_workbook(source_file, data_only=True)
    ws_src = wb_src["AKL"]

    data = ws_src.values
    columns = next(data)
    df_src = pd.DataFrame(data, columns=columns)

    wb_dst = load_workbook(target_path)

    if "ItemStockList" not in wb_dst.sheetnames:
        ws_dst = wb_dst.create_sheet(title="ItemStockList")
    else:
        ws_dst = wb_dst["ItemStockList"]

        max_row = ws_dst.max_row
        if max_row > 1:
            ws_dst.delete_rows(2, max_row - 1)

    for i, row in enumerate(df_src.itertuples(index=False), start=2):
        for j, value in enumerate(row, start=1):
            ws_dst.cell(row=i, column=j, value=value)

    wb_dst.save(target_path)
    wb_src.close()
    wb_dst.close()

    messagebox.showinfo("Created", f"✅ Updated data in '{new_filename}'")


def exit_program():
    root.destroy()


root = Tk()
root.title("📦 Update Item Stock List")
root.geometry("400x120")

Button(root, text="Update", bg="#4CAF50", fg="white",
       font=("Arial", 11), width=25, command=create_po_file).pack(pady=15)

Button(root, text="Exit", bg="#d9534f", fg="white",
       font=("Arial", 11), width=25, command=exit_program).pack(pady=5)

root.mainloop()
